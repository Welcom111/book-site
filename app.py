from flask import Flask, render_template, request, redirect, url_for, session, send_file, flash, jsonify
from flask import send_from_directory
from functools import wraps
from datetime import date, datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
import hashlib
import json
import os
from pathlib import Path
import pprint
import re
import secrets
from threading import Lock

import openpyxl
import requests
from werkzeug.security import check_password_hash, generate_password_hash

from local_config import ACCOUNTS, API_KEY, SECRET_KEY

EXCEL_FILE = 'books.xlsx'
ACCOUNTS_SHEET = '__accounts__'
EXCEL_LOCK = Lock()
LOGIN_RE = re.compile(r'^[A-Za-zА-Яа-яЁё0-9_-]{3,24}$')
BOOK_HEADERS = [
    'Название', 'id', 'Автор', 'Год', 'Страниц', 'Рейтинг', 'Жанры',
    'Добавил', 'Картинка', 'Избранное', 'Дата начала прочтения',
    'Дата окончания прочтения', 'Рейтинг', 'Романтично', 'Дружба',
    'Юмор', 'Стекло', 'Сюжет', 'Химия и напряжение', 'Мои мысли',
    'Популярные тропы',
]
ACCOUNT_HEADERS = ['Логин', 'Аватар', 'Публичная библиотека']
JOURNAL_COLUMNS = {
    'name': 1, 'author': 3, 'pages': 5, 'categories': 7,
    'date_started': 11, 'date_finished': 12, 'journal_rating': 13,
    'romance': 14, 'friendship': 15, 'humor': 16, 'heartbreak': 17,
    'plot': 18, 'chemistry': 19, 'thoughts': 20, 'tropes': 21,
}

app = Flask(__name__)
app.secret_key = SECRET_KEY
app.config['MAX_CONTENT_LENGTH'] = 8 * 1024 * 1024


def init_excel():
    if os.path.exists(EXCEL_FILE):
        return
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'library'
    sheet.append(BOOK_HEADERS)
    workbook.save(EXCEL_FILE)


init_excel()


def ensure_accounts_sheet(workbook):
    if ACCOUNTS_SHEET not in workbook.sheetnames:
        sheet = workbook.create_sheet(ACCOUNTS_SHEET)
        sheet.append(ACCOUNT_HEADERS)
        sheet.sheet_state = 'hidden'
    sheet = workbook[ACCOUNTS_SHEET]
    if [sheet.cell(row=1, column=column).value for column in range(1, 5)] == ['Логин', 'Хеш пароля', 'Аватар', 'Публичная библиотека']:
        profiles = [(row[0], row[2], row[3]) for row in sheet.iter_rows(min_row=2, values_only=True) if row[0]]
        sheet.delete_rows(1, sheet.max_row)
        sheet.append(ACCOUNT_HEADERS)
        for profile in profiles:
            sheet.append(profile)
    return sheet


def find_account(workbook, login):
    if ACCOUNTS_SHEET not in workbook.sheetnames:
        return None
    for row in workbook[ACCOUNTS_SHEET].iter_rows(min_row=2, values_only=True):
        if row[0] and str(row[0]).casefold() == login.casefold():
            return {
                'login': str(row[0]),
                'avatar': row[1] or '',
                'is_public': str(row[2]).lower() in ('1', 'true', 'да', 'yes'),
            }
    return None


def configured_login(login):
    return next((item for item in ACCOUNTS if item.casefold() == login.casefold()), None)


def verify_account_password(login, password):
    """Checks a password stored as a Werkzeug password hash."""
    stored_password = ACCOUNTS.get(login, '')
    return check_password_hash(stored_password, password)


def persist_accounts():
    config_path = Path(__file__).with_name('local_config.py')
    content = config_path.read_text(encoding='utf-8')
    account_block = '# Accounts managed by app\nACCOUNTS = ' + pprint.pformat(ACCOUNTS, sort_dicts=True, width=100) + '\n# End managed accounts'
    updated, replacements = re.subn(
        r'(?s)# Accounts managed by app\nACCOUNTS = .*?\n# End managed accounts',
        account_block,
        content,
    )
    if replacements != 1:
        raise RuntimeError('Не найден блок ACCOUNTS в local_config.py')
    config_path.write_text(updated, encoding='utf-8')


def migrate_legacy_passwords():
    """Converts pre-existing clear-text passwords to hashes on startup."""
    changed = False
    for login, stored_password in list(ACCOUNTS.items()):
        if not stored_password.startswith(('scrypt:', 'pbkdf2:')):
            ACCOUNTS[login] = generate_password_hash(stored_password)
            changed = True
    if changed:
        persist_accounts()


migrate_legacy_passwords()


def account_sheet(workbook, login):
    account = find_account(workbook, login)
    if not account or account['login'] not in workbook.sheetnames:
        return None
    return workbook[account['login']]


def require_login(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get('login'):
            flash('Войдите в аккаунт, чтобы открыть свою библиотеку', 'error')
            return redirect(url_for('login'))
        return view(*args, **kwargs)
    return wrapped


@app.context_processor
def inject_current_user():
    return {
        'current_login': session.get('login'),
        'current_avatar': session.get('avatar', ''),
    }


def make_cover_filename(book_id, extension='jpg'):
    safe_id = re.sub(r'[\\/*?:"<>|.\s]', '_', str(book_id or ''))
    return f'{safe_id}.{extension}'


def find_cover_image(book_id):
    for extension in ('jpg', 'jpeg', 'png', 'webp'):
        filename = make_cover_filename(book_id, extension)
        if os.path.exists(os.path.join('static', 'images', filename)):
            return f'images/{filename}'
    return None


def make_manual_book_id(book_name):
    normalized_name = re.sub(r'\s+', ' ', book_name.strip()).casefold()
    name_hash = hashlib.sha256(normalized_name.encode('utf-8')).hexdigest()[:8]
    return f'manual-{name_hash}-{secrets.token_hex(3)}'


def image_to_jpeg(uploaded_file, output_path):
    if not uploaded_file or not uploaded_file.filename:
        return None
    extension = uploaded_file.filename.rsplit('.', 1)[-1].lower() if '.' in uploaded_file.filename else ''
    if extension not in {'jpg', 'jpeg', 'png', 'webp'}:
        raise ValueError('Допустимы только изображения JPG, PNG или WebP')

    from PIL import Image as PILImage
    from PIL import ImageOps

    try:
        image = PILImage.open(uploaded_file.stream)
        image.verify()
        uploaded_file.stream.seek(0)
        image = ImageOps.exif_transpose(PILImage.open(uploaded_file.stream))
    except Exception as error:
        raise ValueError('Не удалось прочитать загруженное изображение') from error

    if image.mode in ('RGBA', 'P', 'LA'):
        background = PILImage.new('RGB', image.size, (255, 255, 255))
        if image.mode == 'P':
            image = image.convert('RGBA')
        background.paste(image, mask=image.getchannel('A') if image.mode in ('RGBA', 'LA') else None)
        image = background
    elif image.mode != 'RGB':
        image = image.convert('RGB')
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    image.save(output_path, 'JPEG', quality=95, optimize=True)


def save_uploaded_cover(uploaded_file, book_id):
    filename = make_cover_filename(book_id)
    image_to_jpeg(uploaded_file, os.path.join('static', 'images', filename))
    return f'images/{filename}'


def save_uploaded_avatar(uploaded_file, login):
    safe_login = re.sub(r'[^A-Za-zА-Яа-яЁё0-9_-]', '_', login)
    filename = f'avatar_{safe_login}.jpg'
    image_to_jpeg(uploaded_file, os.path.join('static', 'avatars', filename))
    return f'avatars/{filename}'


def download_and_insert_image(sheet, row_num, image_url, book_id):
    try:
        import io
        from PIL import Image as PILImage

        response = requests.get(image_url, headers={'User-Agent': 'Mozilla/5.0'}, timeout=(10, 60))
        response.raise_for_status()
        image = PILImage.open(io.BytesIO(response.content)).convert('RGB')
        filename = make_cover_filename(book_id)
        path = os.path.join('static', 'images', filename)
        os.makedirs(os.path.dirname(path), exist_ok=True)
        image.save(path, 'JPEG', quality=95, optimize=True)
        excel_image = XLImage(path)
        excel_image.width, excel_image.height = 60, 80
        sheet.add_image(excel_image, f'I{row_num}')
        sheet.row_dimensions[row_num].height = 90
        sheet.column_dimensions['I'].width = 12
        return f'images/{filename}'
    except Exception as error:
        print(f'Ошибка при загрузке обложки: {error}')
        return None


def books_from_sheet(sheet):
    result = []
    for row_index in range(2, sheet.max_row + 1):
        name = sheet.cell(row=row_index, column=1).value
        if not name:
            continue
        book_id = sheet.cell(row=row_index, column=2).value
        result.append({
            'row_index': row_index,
            'name': name,
            'id': book_id,
            'author': sheet.cell(row=row_index, column=3).value,
            'year': sheet.cell(row=row_index, column=4).value,
            'pages': sheet.cell(row=row_index, column=5).value,
            'rating': sheet.cell(row=row_index, column=6).value,
            'categories': sheet.cell(row=row_index, column=7).value,
            'favorite': str(sheet.cell(row=row_index, column=10).value or ''),
            'image_path': find_cover_image(book_id),
        })
    return result


def get_current_library():
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = account_sheet(workbook, session['login'])
    if sheet is None:
        workbook.close()
        session.clear()
        return None, None
    return workbook, sheet


def format_excel_date(value):
    if isinstance(value, (date, datetime)):
        return value.strftime('%Y-%m-%d')
    return str(value) if value else ''


def parse_form_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, '%Y-%m-%d').date()
    except ValueError:
        return None


def parse_score(value):
    try:
        return max(0, min(5, int(value))) if value not in (None, '') else None
    except (TypeError, ValueError):
        return None


def parse_non_negative_int(value):
    try:
        return max(0, int(value)) if value not in (None, '') else None
    except (TypeError, ValueError):
        return None


def categories_from_sheet(sheet):
    categories = set()
    for row in sheet.iter_rows(min_row=2, min_col=7, max_col=7, values_only=True):
        if row[0]:
            categories.update(item.strip() for item in str(row[0]).split(';') if item.strip())
    return sorted(categories)


def authors_from_sheet(sheet):
    authors = set()
    for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3, values_only=True):
        if row[0]:
            authors.update(item.strip() for item in str(row[0]).split(',') if item.strip())
    return sorted(authors)


@app.route('/favicon.ico')
def favicon():
    return send_from_directory('static', 'favicon.ico')


@app.route('/')
def index():
    if session.get('login'):
        return redirect(url_for('home'))
    return render_template('landing.html')


@app.route('/register', methods=['GET', 'POST'])
def register():
    if session.get('login'):
        return redirect(url_for('home'))
    if request.method == 'POST':
        login = request.form.get('login', '').strip()
        password = request.form.get('password', '')
        password_repeat = request.form.get('password_repeat', '')
        if not LOGIN_RE.fullmatch(login):
            flash('Логин: от 3 до 24 символов — буквы, цифры, дефис или подчёркивание', 'error')
        elif len(password) < 6:
            flash('Пароль должен содержать минимум 6 символов', 'error')
        elif password != password_repeat:
            flash('Пароли не совпадают', 'error')
        else:
            with EXCEL_LOCK:
                workbook = openpyxl.load_workbook(EXCEL_FILE)
                accounts = ensure_accounts_sheet(workbook)
                if configured_login(login) or login in workbook.sheetnames:
                    workbook.close()
                    flash('Такой логин уже занят', 'error')
                else:
                    sheet = workbook.create_sheet(login)
                    sheet.append(BOOK_HEADERS)
                    accounts.append([login, '', False])
                    workbook.save(EXCEL_FILE)
                    workbook.close()
                    ACCOUNTS[login] = generate_password_hash(password)
                    persist_accounts()
                    session['login'] = login
                    session['avatar'] = ''
                    flash('Аккаунт создан. Добро пожаловать!', 'success')
                    return redirect(url_for('home'))
    return render_template('auth.html', mode='register')


@app.route('/login', methods=['GET', 'POST'])
def login():
    if session.get('login'):
        return redirect(url_for('home'))
    if request.method == 'POST':
        login_value = request.form.get('login', '').strip()
        password = request.form.get('password', '')
        login_name = configured_login(login_value)
        workbook = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
        account = find_account(workbook, login_name) if login_name else None
        workbook.close()
        if account and verify_account_password(login_name, password):
            session['login'] = account['login']
            session['avatar'] = account['avatar']
            flash('Вы вошли в свою библиотеку', 'success')
            return redirect(url_for('home'))
        flash('Неверный логин или пароль', 'error')
    return render_template('auth.html', mode='login')


@app.route('/logout', methods=['POST'])
@require_login
def logout():
    session.clear()
    flash('Вы вышли из аккаунта', 'success')
    return redirect(url_for('index'))


@app.route('/home')
@require_login
def home():
    workbook, sheet = get_current_library()
    if sheet is None:
        return redirect(url_for('login'))
    books = books_from_sheet(sheet)
    workbook.close()
    return render_template('index.html', books_count=len(books), featured_books=[book for book in books if book['image_path']][:2])


@app.route('/books')
@require_login
def books():
    workbook, sheet = get_current_library()
    if sheet is None:
        return redirect(url_for('login'))
    books_list = books_from_sheet(sheet)
    workbook.close()
    return render_template('books.html', books=books_list)


@app.route('/books/<book_id>', methods=['GET', 'POST'])
@require_login
def book_journal(book_id):
    with EXCEL_LOCK:
        workbook, sheet = get_current_library()
        if sheet is None:
            return redirect(url_for('login'))
        row_index = next((row for row in range(2, sheet.max_row + 1) if str(sheet.cell(row=row, column=2).value or '') == str(book_id)), None)
        if row_index is None:
            workbook.close()
            flash('Книга не найдена в вашей библиотеке', 'error')
            return redirect(url_for('books'))

        if request.method == 'POST':
            book_name = request.form.get('name', '').strip()
            cover = request.files.get('cover')
            try:
                cover_path = save_uploaded_cover(cover, book_id) if cover and cover.filename else None
            except ValueError as error:
                workbook.close()
                flash(str(error), 'error')
                return redirect(url_for('book_journal', book_id=book_id))
            sheet.cell(row=row_index, column=1, value=book_name)
            sheet.cell(row=row_index, column=3, value=request.form.get('author', '').strip())
            sheet.cell(row=row_index, column=5, value=parse_non_negative_int(request.form.get('pages')))
            sheet.cell(row=row_index, column=7, value=request.form.get('categories', '').strip())
            sheet.cell(row=row_index, column=11, value=parse_form_date(request.form.get('date_started')))
            sheet.cell(row=row_index, column=12, value=parse_form_date(request.form.get('date_finished')))
            for field in ('journal_rating', 'romance', 'friendship', 'humor', 'heartbreak', 'plot', 'chemistry'):
                sheet.cell(row=row_index, column=JOURNAL_COLUMNS[field], value=parse_score(request.form.get(field)))
            sheet.cell(row=row_index, column=20, value=request.form.get('thoughts', '').strip())
            sheet.cell(row=row_index, column=21, value=request.form.get('tropes', '').strip())
            if cover_path:
                sheet.cell(row=row_index, column=9, value=cover_path)
            workbook.save(EXCEL_FILE)
            workbook.close()
            flash('Дневник книги сохранён', 'success')
            return redirect(url_for('books'))

        journal = {field: sheet.cell(row=row_index, column=column).value for field, column in JOURNAL_COLUMNS.items()}
        journal['date_started'] = format_excel_date(journal['date_started'])
        journal['date_finished'] = format_excel_date(journal['date_finished'])
        journal['image_path'] = find_cover_image(book_id)
        workbook.close()
    return render_template('book_journal.html', book=journal, book_id=book_id)


@app.route('/add_book')
@require_login
def add_book():
    return render_template('add_book.html')


@app.route('/add_book_manual', methods=['POST'])
@require_login
def add_book_manual():
    title = request.form.get('manual_title', '').strip()
    cover = request.files.get('manual_cover')
    if not title:
        flash('Укажите название книги', 'error')
        return redirect(url_for('add_book'))
    with EXCEL_LOCK:
        workbook, sheet = get_current_library()
        if sheet is None:
            return redirect(url_for('login'))
        existing_ids = {str(sheet.cell(row=row, column=2).value or '') for row in range(2, sheet.max_row + 1)}
        book_id = make_manual_book_id(title)
        while book_id in existing_ids:
            book_id = make_manual_book_id(title)
        try:
            cover_path = save_uploaded_cover(cover, book_id) if cover and cover.filename else ''
        except ValueError as error:
            workbook.close()
            flash(str(error), 'error')
            return redirect(url_for('add_book'))
        sheet.append([title, book_id, '', '', '', '', '', session['login'], cover_path, ''])
        workbook.save(EXCEL_FILE)
        workbook.close()
    flash(f"Книга '{title}' успешно добавлена!", 'success')
    return redirect(url_for('books'))


@app.route('/search_books', methods=['POST'])
@require_login
def search_books():
    search_text = request.form.get('book_name', '').strip()
    if not search_text:
        return redirect(url_for('add_book'))
    try:
        response = requests.get(
            'https://www.googleapis.com/books/v1/volumes',
            params={'q': search_text, 'key': API_KEY, 'langRestrict': 'ru', 'maxResults': 20},
            headers={'User-Agent': 'Mozilla/5.0'}, timeout=10,
        )
        response.raise_for_status()
        results = []
        for item in response.json().get('items', []):
            info = item.get('volumeInfo', {})
            results.append({
                'name': info.get('title', 'Без названия'),
                'id': item.get('id'),
                'authors': ', '.join(info.get('authors', ['Автор не указан'])),
                'year': (info.get('publishedDate') or '')[:4],
                'cover_url': (info.get('imageLinks') or {}).get('thumbnail', ''),
            })
        return render_template('search_results.html', results=results, search_text=search_text)
    except requests.RequestException:
        flash('Не удалось получить данные из Google Books', 'error')
        return redirect(url_for('add_book'))


@app.route('/add_book_by_id', methods=['POST'])
@require_login
def add_book_by_id():
    book_id = request.form.get('book_id', '').strip()
    if not book_id:
        flash('Укажите Google Books ID', 'error')
        return redirect(url_for('add_book'))
    try:
        response = requests.get(f'https://www.googleapis.com/books/v1/volumes/{book_id}', params={'key': API_KEY}, timeout=10)
        response.raise_for_status()
        info = response.json().get('volumeInfo', {})
    except requests.RequestException:
        flash('Не удалось получить книгу из Google Books', 'error')
        return redirect(url_for('add_book'))

    with EXCEL_LOCK:
        workbook, sheet = get_current_library()
        if sheet is None:
            return redirect(url_for('login'))
        if any(str(sheet.cell(row=row, column=2).value or '') == book_id for row in range(2, sheet.max_row + 1)):
            workbook.close()
            flash('Эта книга уже есть в вашей библиотеке', 'error')
            return redirect(url_for('books'))
        next_row = sheet.max_row + 1
        title = info.get('title', 'Без названия')
        sheet.append([
            title, book_id, ', '.join(info.get('authors', ['Автор не указан'])),
            (info.get('publishedDate') or '')[:4], info.get('pageCount', ''),
            info.get('averageRating', ''), '; '.join(info.get('categories', [])),
            session['login'], '', '',
        ])
        image_url = (info.get('imageLinks') or {}).get('medium') or (info.get('imageLinks') or {}).get('thumbnail')
        if image_url:
            image_path = download_and_insert_image(sheet, next_row, image_url, book_id)
            if image_path:
                sheet.cell(row=next_row, column=9, value=image_path)
        workbook.save(EXCEL_FILE)
        workbook.close()
    flash(f"Книга '{title}' успешно добавлена!", 'success')
    return redirect(url_for('books'))


@app.route('/filter')
@require_login
def filter_books():
    return render_template('filter.html')


@app.route('/filter/author', methods=['GET', 'POST'])
@require_login
def filter_by_author():
    workbook, sheet = get_current_library()
    if sheet is None:
        return redirect(url_for('login'))
    authors = authors_from_sheet(sheet)
    if request.method == 'POST':
        author = request.form.get('author', '')
        matches = [row[2] for row in sheet.iter_rows(min_row=2, values_only=True) if row[2] and author.casefold() in str(row[2]).casefold()]
        workbook.close()
        return render_template('filter_results.html', games=matches, filter_text=f"по автору '{author}'")
    workbook.close()
    return render_template('filter_author.html', authors=authors)


@app.route('/filter/category', methods=['GET', 'POST'])
@require_login
def filter_by_category():
    workbook, sheet = get_current_library()
    if sheet is None:
        return redirect(url_for('login'))
    categories = categories_from_sheet(sheet)
    if request.method == 'POST':
        category = request.form.get('category', '')
        matches = [row[0] for row in sheet.iter_rows(min_row=2, values_only=True) if row[6] and category in str(row[6]).split(';')]
        workbook.close()
        return render_template('filter_results.html', games=matches, filter_text=f"в жанре '{category}'")
    workbook.close()
    return render_template('filter_category.html', categories=categories)


@app.route('/export_excel')
@require_login
def export_excel():
    workbook, sheet = get_current_library()
    if sheet is None:
        return redirect(url_for('login'))
    export = Workbook()
    export_sheet = export.active
    export_sheet.title = session['login']
    for row in sheet.iter_rows(values_only=True):
        export_sheet.append(list(row))
    workbook.close()
    output = BytesIO()
    export.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=f"{session['login']}_books.xlsx")


@app.route('/settings', methods=['GET', 'POST'])
@require_login
def settings():
    with EXCEL_LOCK:
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        account = find_account(workbook, session['login'])
        if not account:
            workbook.close()
            session.clear()
            return redirect(url_for('login'))
        if request.method == 'POST':
            accounts = workbook[ACCOUNTS_SHEET]
            row_number = next(row for row in range(2, accounts.max_row + 1) if str(accounts.cell(row=row, column=1).value).casefold() == account['login'].casefold())
            new_password = request.form.get('new_password', '')
            password_repeat = request.form.get('password_repeat', '')
            if new_password and (len(new_password) < 6 or new_password != password_repeat):
                workbook.close()
                flash('Новый пароль должен содержать минимум 6 символов, и поля должны совпадать', 'error')
                return redirect(url_for('settings'))
            try:
                avatar = request.files.get('avatar')
                avatar_path = save_uploaded_avatar(avatar, account['login']) if avatar and avatar.filename else account['avatar']
            except ValueError as error:
                workbook.close()
                flash(str(error), 'error')
                return redirect(url_for('settings'))
            accounts.cell(row=row_number, column=2, value=avatar_path)
            accounts.cell(row=row_number, column=3, value='да' if request.form.get('is_public') else 'нет')
            if new_password:
                ACCOUNTS[account['login']] = generate_password_hash(new_password)
            workbook.save(EXCEL_FILE)
            workbook.close()
            if new_password:
                persist_accounts()
            session['avatar'] = avatar_path
            flash('Настройки сохранены', 'success')
            return redirect(url_for('settings'))
        workbook.close()
    return render_template('settings.html', account=account)


@app.route('/libraries')
def public_libraries():
    workbook = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
    libraries = []
    if ACCOUNTS_SHEET in workbook.sheetnames:
        for row in workbook[ACCOUNTS_SHEET].iter_rows(min_row=2, values_only=True):
            if row[0] and configured_login(str(row[0])) and str(row[2]).lower() in ('1', 'true', 'да', 'yes'):
                libraries.append({'login': row[0], 'avatar': row[1] or ''})
    workbook.close()
    return render_template('public_libraries.html', libraries=libraries)


@app.route('/libraries/<login>')
def public_library(login):
    workbook = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
    account = find_account(workbook, login)
    if not account or not account['is_public'] or account['login'] not in workbook.sheetnames:
        workbook.close()
        flash('Эта библиотека скрыта или не найдена', 'error')
        return redirect(url_for('public_libraries'))
    books_list = books_from_sheet(workbook[account['login']])
    workbook.close()
    return render_template('public_books.html', books=books_list, owner=account)


@app.route('/toggle_favorite', methods=['POST'])
@require_login
def toggle_favorite():
    data = request.get_json(silent=True) or {}
    book_id = data.get('book_id')
    if not book_id:
        return jsonify({'success': False, 'error': 'Не указан ID книги'}), 400
    with EXCEL_LOCK:
        workbook, sheet = get_current_library()
        if sheet is None:
            return jsonify({'success': False, 'error': 'Требуется вход'}), 401
        for row in range(2, sheet.max_row + 1):
            if str(sheet.cell(row=row, column=2).value or '') == str(book_id):
                current = str(sheet.cell(row=row, column=10).value or '').lower()
                value = 'нет' if current == 'да' else 'да'
                sheet.cell(row=row, column=10, value=value)
                workbook.save(EXCEL_FILE)
                workbook.close()
                return jsonify({'success': True, 'favorite': value})
        workbook.close()
    return jsonify({'success': False, 'error': 'Книга не найдена'}), 404


@app.route('/static/<path:filename>')
def serve_static(filename):
    return send_from_directory('static', filename)


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5001, debug=True)
